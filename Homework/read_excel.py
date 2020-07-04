# @Time    : 2020/6/30 14:57
# @Author  : 叉叉
# @email   : lovechachaf@163.com
from openpyxl import load_workbook
def read_excel(file_name,sheet_name):
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]
    test_case=[]
    for i in range(2,sheet.max_row+1):
        case=[]
        for j in range(3,sheet.max_column+1):
            value=sheet.cell(row=i,column=j).value
            case.append(value)
        test_case.append(case)
    return test_case

def wrtte_excel(file_name,sheet_name,row,column,result):
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    sheet.cell(row=row,column=column).value=result
    wb.save()
