# @Time    : 2020/6/30 15:17
# @Author  : 叉叉
# @email   : lovechachaf@163.com
from Homework import http_request
from Homework import read_excel
from openpyxl import load_workbook
import json
test_case=read_excel.read_excel('recharge_testcase.xlsx','recharge')
login_case=test_case[0]
ip='http://120.78.128.25:8766'
login_result=http_request.http_request(ip+login_case[1],eval(login_case[2]),None,login_case[0])
token=login_result['data']['token_info']['token']
recharge_res=[]
for i in range(1,len(test_case)):
    recharge=http_request.http_request(ip+test_case[i][1],eval(test_case[i][2]),'Bearer '+token,test_case[i][0])
    recharge_res.append(recharge)
wb=load_workbook('recharge_testcase.xlsx')
sheet=wb['recharge']
for res in range(2,len(recharge_res)+2):
    sheet.cell(row=res+1,column=8).value=json.dumps(recharge_res[res-2])
    wb.save('recharge_testcase.xlsx')


